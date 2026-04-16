import io
from datetime import date, timedelta
from decimal import Decimal

from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Count, Q
from django.utils import timezone

from orders.models import Order, OrderProduct


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def _get_date_range(period, date_from_str, date_to_str):
    today = date.today()
    if period == 'daily':
        return today, today
    elif period == 'weekly':
        start = today - timedelta(days=today.weekday())   # Monday
        return start, today
    elif period == 'monthly':
        return today.replace(day=1), today
    elif period == 'yearly':
        return today.replace(month=1, day=1), today
    else:  # custom
        try:
            d_from = date.fromisoformat(date_from_str)
            d_to   = date.fromisoformat(date_to_str)
            if d_from > d_to:
                d_from, d_to = d_to, d_from
        except (ValueError, TypeError):
            d_from = today - timedelta(days=30)
            d_to   = today
        return d_from, d_to


def _build_report(date_from, date_to):
    """Returns queryset + aggregate summary for the date range."""
    qs = Order.objects.filter(
        is_ordered=True,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).select_related('user', 'payment').order_by('-created_at')

    summary = qs.aggregate(
        total_orders    = Count('id'),
        total_revenue   = Sum('order_total'),
        total_discount  = Sum('discount'),
        total_wallet    = Sum('wallet_used'),
    )
    # Fill None with 0
    for k in summary:
        if summary[k] is None:
            summary[k] = Decimal('0')

    # Cancelled/Returned count
    summary['cancelled_count'] = qs.filter(status='Cancelled').count()
    summary['returned_count']  = qs.filter(status__in=['Return Requested', 'Returned']).count()

    return qs, summary


# ─────────────────────────────────────────────────────────
# SALES REPORT PAGE
# ─────────────────────────────────────────────────────────
@staff_member_required(login_url='admin_login')
def admin_sales_report(request):
    period      = request.GET.get('period', 'monthly')
    date_from_s = request.GET.get('date_from', '')
    date_to_s   = request.GET.get('date_to', '')

    date_from, date_to = _get_date_range(period, date_from_s, date_to_s)
    orders, summary    = _build_report(date_from, date_to)

    # Pagination
    from django.core.paginator import Paginator
    paginator  = Paginator(orders, 20)
    page_obj   = paginator.get_page(request.GET.get('page', 1))

    PERIODS = [
        ('daily',   'Today'),
        ('weekly',  'This Week'),
        ('monthly', 'This Month'),
        ('yearly',  'This Year'),
        ('custom',  'Custom'),
    ]
    context = {
        'page_obj'  : page_obj,
        'summary'   : summary,
        'period'    : period,
        'date_from' : date_from,
        'date_to'   : date_to,
        'periods'   : PERIODS,
    }
    return render(request, 'adminpanel/admin_sales_report.html', context)


# ─────────────────────────────────────────────────────────
# DOWNLOAD PDF
# ─────────────────────────────────────────────────────────
@staff_member_required(login_url='admin_login')
def admin_sales_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.units import cm

    period      = request.GET.get('period', 'monthly')
    date_from_s = request.GET.get('date_from', '')
    date_to_s   = request.GET.get('date_to', '')
    date_from, date_to = _get_date_range(period, date_from_s, date_to_s)
    orders, summary    = _build_report(date_from, date_to)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles  = getSampleStyleSheet()
    content = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontSize=16, spaceAfter=4)
    content.append(Paragraph('Orbit Watch Collection — Sales Report', title_style))
    content.append(Paragraph(
        f'Period: {date_from.strftime("%d %b %Y")} to {date_to.strftime("%d %b %Y")}',
        styles['Normal']
    ))
    content.append(Spacer(1, 0.4*cm))

    # Summary table
    summary_data = [
        ['Total Orders', 'Total Revenue', 'Total Discount', 'Wallet Used', 'Cancelled', 'Returned'],
        [
            str(summary['total_orders']),
            f"Rs.{summary['total_revenue']}",
            f"Rs.{summary['total_discount']}",
            f"Rs.{summary['total_wallet']}",
            str(summary['cancelled_count']),
            str(summary['returned_count']),
        ]
    ]
    s_table = Table(summary_data, colWidths=[3.5*cm]*6)
    s_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3167eb')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f0f4ff'), colors.white]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#c0c0c0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    content.append(s_table)
    content.append(Spacer(1, 0.5*cm))

    # Orders table
    headers = ['Order #', 'Date', 'Customer', 'Payment', 'Coupon Disc.', 'Wallet', 'Total', 'Status']
    rows    = [headers]
    for o in orders:
        rows.append([
            o.order_number,
            o.created_at.strftime('%d %b %Y'),
            o.user.email if o.user else '—',
            o.payment.payment_method if o.payment else '—',
            f"Rs.{o.discount}",
            f"Rs.{o.wallet_used}",
            f"Rs.{o.order_total}",
            o.status,
        ])

    col_w = [3*cm, 2.5*cm, 5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm]
    o_table = Table(rows, colWidths=col_w, repeatRows=1)
    o_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c2c2c')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#d0d0d0')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',      (2, 1), (2, -1), 'LEFT'),   # email left-aligned
    ]))
    content.append(o_table)

    doc.build(content)
    buffer.seek(0)
    filename = f"orbit_sales_{date_from}_{date_to}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────
# DOWNLOAD EXCEL
# ─────────────────────────────────────────────────────────
@staff_member_required(login_url='admin_login')
def admin_sales_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            'openpyxl is not installed. Run: pip install openpyxl',
            content_type='text/plain', status=500
        )

    period      = request.GET.get('period', 'monthly')
    date_from_s = request.GET.get('date_from', '')
    date_to_s   = request.GET.get('date_to', '')
    date_from, date_to = _get_date_range(period, date_from_s, date_to_s)
    orders, summary    = _build_report(date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    # ── Styles ──────────────────────────────────────────
    BLUE      = '3167EB'
    DARK      = '2C2C2C'
    LIGHT_BG  = 'F0F4FF'
    HDR_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    BODY_FONT = Font(name='Calibri', size=10)
    CENTER    = Alignment(horizontal='center', vertical='center')
    LEFT      = Alignment(horizontal='left',   vertical='center')
    thin      = Side(style='thin', color='C0C0C0')
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, value, bg=DARK):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = HDR_FONT
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = CENTER
        c.border    = BORDER
        return c

    def body_cell(ws, row, col, value, align=CENTER, bg=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = BODY_FONT
        c.alignment = align
        c.border    = BORDER
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        return c

    # ── Title ────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value     = f'Orbit Watch Collection — Sales Report'
    title.font      = Font(name='Calibri', bold=True, size=14, color=DARK)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    sub = ws['A2']
    sub.value     = f'Period: {date_from.strftime("%d %b %Y")}  to  {date_to.strftime("%d %b %Y")}'
    sub.font      = Font(name='Calibri', size=11, color='555555')
    sub.alignment = CENTER
    ws.row_dimensions[2].height = 20

    # ── Summary row ──────────────────────────────────────
    ws.row_dimensions[4].height = 22
    for col, h in enumerate(['Total Orders', 'Total Revenue', 'Coupon Discounts',
                              'Wallet Used', 'Cancelled', 'Returned'], start=1):
        hdr_cell(ws, 4, col, h, bg=BLUE)

    ws.row_dimensions[5].height = 20
    for col, v in enumerate([
        summary['total_orders'],
        f"Rs.{summary['total_revenue']}",
        f"Rs.{summary['total_discount']}",
        f"Rs.{summary['total_wallet']}",
        summary['cancelled_count'],
        summary['returned_count'],
    ], start=1):
        body_cell(ws, 5, col, v, bg=LIGHT_BG)

    # ── Orders header ─────────────────────────────────────
    ws.row_dimensions[7].height = 22
    headers = ['Order #', 'Date', 'Customer Email', 'Payment Method',
               'Coupon Discount', 'Wallet Used', 'Order Total', 'Status']
    for col, h in enumerate(headers, start=1):
        hdr_cell(ws, 7, col, h, bg=DARK)

    # ── Orders rows ───────────────────────────────────────
    for row_idx, o in enumerate(orders, start=8):
        bg = LIGHT_BG if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 18
        body_cell(ws, row_idx, 1, o.order_number, bg=bg)
        body_cell(ws, row_idx, 2, o.created_at.strftime('%d %b %Y'), bg=bg)
        body_cell(ws, row_idx, 3, o.user.email if o.user else '—', align=LEFT, bg=bg)
        body_cell(ws, row_idx, 4, o.payment.payment_method if o.payment else '—', bg=bg)
        body_cell(ws, row_idx, 5, float(o.discount), bg=bg)
        body_cell(ws, row_idx, 6, float(o.wallet_used), bg=bg)
        body_cell(ws, row_idx, 7, float(o.order_total), bg=bg)
        body_cell(ws, row_idx, 8, o.status, bg=bg)

    # ── Column widths ─────────────────────────────────────
    for col, w in enumerate([16, 14, 30, 18, 16, 14, 14, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Save ──────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"orbit_sales_{date_from}_{date_to}.xlsx"
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response